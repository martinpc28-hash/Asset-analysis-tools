"""
informe_html.py
===============
Genera el INFORME HTML interactivo (y un Excel de metricas) a partir de
cualquier Excel que ya tenga las series y el benchmark.

Reconoce automaticamente TRES formatos:

  A) PLANTILLA Comparativa (la nueva, generica)
     - Hoja 'Datos', filas desde la 3
     - Columnas: A=Fecha, B=Activo 1, C=Activo 2 (nombres editables en B2/C2)

  B) PLANTILLA vs SP500 (la antigua, sigue funcionando)
     - Hoja 'Datos', columnas: A=Fecha, B=Mi Serie, C=SP500

  C) REPORTE generado por analizador.py
     - Hoja 'Precios_alineados'
     - Columnas: Fecha + 1 o mas series + benchmark

No descarga nada de internet. Funciona aunque Yahoo este bloqueado.

Uso:
    python informe_html.py
    python informe_html.py "Plantilla_Analisis_Comparativa.xlsx"
    python informe_html.py mi_reporte_analizador.xlsx --rf 0.05
"""
import sys
import argparse
from pathlib import Path
from datetime import datetime

import pandas as pd

from analizador import calcular_metricas, exportar_html, exportar_excel


def _autodetectar_y_leer(ruta: str) -> tuple[pd.DataFrame, str]:
    """Devuelve (df_combinado, nombre_columna_benchmark)."""
    from openpyxl import load_workbook
    wb = load_workbook(ruta, data_only=True, read_only=True)
    hojas = set(wb.sheetnames)

    # CASO C: reporte de analizador.py
    if "Precios_alineados" in hojas:
        print("  -> Detectado: REPORTE de analizador.py (hoja Precios_alineados)")
        df = pd.read_excel(ruta, sheet_name="Precios_alineados")
        df.columns = [str(c).strip() for c in df.columns]
        fecha_col = df.columns[0]
        df[fecha_col] = pd.to_datetime(df[fecha_col], errors="coerce")
        df = df.dropna(subset=[fecha_col]).set_index(fecha_col)
        df.index.name = "Fecha"
        for c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna(how="any").sort_index()
        # benchmark = columna que parezca indice o la ultima
        candidatos = [c for c in df.columns
                      if c.upper() in ("SP500", "^GSPC", "^GSPM", "SPY", "BENCHMARK")]
        bench = candidatos[0] if candidatos else df.columns[-1]
        return df, bench

    # CASO A o B: plantilla (lee headers de B2/C2 dinamicamente)
    sheet = "Datos" if "Datos" in hojas else wb.sheetnames[0]
    ws = wb[sheet]
    nombre_a = ws["B2"].value or "Activo 1"
    nombre_b = ws["C2"].value or "Activo 2"
    print(f"  -> Detectado: PLANTILLA (hoja {sheet})")
    print(f"  -> Activo 1: {nombre_a}    Activo 2 (benchmark): {nombre_b}")

    df = pd.read_excel(ruta, sheet_name=sheet, skiprows=1, usecols="A:C")
    df.columns = ["Fecha", str(nombre_a), str(nombre_b)]
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df = df.dropna(subset=["Fecha"]).set_index("Fecha")
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(how="any").sort_index()
    return df, str(nombre_b)


def leer_rf(ruta: str, default: float = 0.045) -> float:
    """Lee la tasa libre de riesgo de Dashboard!B4 si existe."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, data_only=True, read_only=True)
        if "Dashboard" in wb.sheetnames:
            v = wb["Dashboard"]["B4"].value
            if isinstance(v, (int, float)):
                return float(v)
    except Exception:
        pass
    return default


def main():
    ap = argparse.ArgumentParser(
        description="Genera informe HTML + Excel desde un Excel con datos.")
    ap.add_argument("archivo", nargs="?", default="Plantilla_Analisis_Comparativa.xlsx",
                    help="Ruta al Excel (plantilla o reporte de analizador.py)")
    ap.add_argument("--rf", type=float, default=None,
                    help="Tasa libre de riesgo anual (default: Dashboard!B4 o 4.5%%)")
    args = ap.parse_args()

    ruta = args.archivo

    # Fallback: si pasaron el nombre default y no existe, prueba el viejo
    if not Path(ruta).exists():
        alt = "Plantilla_Analisis_vs_SP500.xlsx"
        if ruta == "Plantilla_Analisis_Comparativa.xlsx" and Path(alt).exists():
            print(f"[info] No encontre '{ruta}', usando '{alt}'")
            ruta = alt
        else:
            print(f"[ERROR] No encuentro el archivo: {ruta}")
            sys.exit(1)

    print(f"\n=== INFORME HTML DESDE EXCEL ===")
    print(f"Leyendo: {ruta}")

    df, bench_col = _autodetectar_y_leer(ruta)
    if df.shape[0] < 30:
        print(f"[AVISO] Solo {df.shape[0]} filas con datos completos. "
              "Resultados poco fiables con tan pocas observaciones.")
        if df.shape[0] < 5:
            print("[ERROR] Insuficientes datos para calcular metricas.")
            sys.exit(1)

    print(f"  -> {df.shape[0]} filas validas: {df.index.min().date()} -> {df.index.max().date()}")
    series_user = [c for c in df.columns if c != bench_col]
    print(f"  -> Benchmark: {bench_col}    Series: {series_user}")

    rf = args.rf if args.rf is not None else leer_rf(ruta)
    print(f"  -> Tasa libre de riesgo: {rf*100:.2f}%")

    resultados = {}
    for s in series_user:
        m = calcular_metricas(df[s], df[bench_col], rf)
        resultados[s] = m
        print(f"  -> {s}: Beta={m['Beta']:.3f}  R2={m['R_cuadrado']:.3f}  "
              f"Sharpe={m['Sharpe_serie']:.3f}  Vol={m['Volatilidad_anual_serie']*100:.1f}%")

    out_dir = Path(ruta).parent
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    base = Path(ruta).stem
    html_out = out_dir / f"{base}_informe_{stamp}.html"
    excel_out = out_dir / f"{base}_metricas_{stamp}.xlsx"

    exportar_html(df, resultados, bench_col, str(html_out))
    exportar_excel(df, resultados, str(excel_out))

    print(f"\n[OK] Generados en {out_dir}:")
    print(f"  - {html_out.name}")
    print(f"  - {excel_out.name}")


if __name__ == "__main__":
    main()