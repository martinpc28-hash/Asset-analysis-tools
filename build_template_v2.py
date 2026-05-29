"""
Construye Plantilla_Analisis_Comparativa.xlsx
=============================================
Plantilla GENERICA: compara cualquier 'Activo 1' vs cualquier 'Activo 2'.
El antiguo SP500 ya no esta hardcoded; la columna C es simplemente
'Activo 2 (benchmark)'.

Estructura:
  Instrucciones    - como usar
  Datos            - Fecha | Activo 1 | Activo 2  (hasta 5000 filas)
  Dashboard        - todas las metricas
  Grafico          - precios base 100 y drawdown
  Calculos         - retornos, excesos, drawdowns, diff (no editar)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule

OUT = "Plantilla_Analisis_Comparativa.xlsx"
N_MAX = 5000
RF_DEFAULT = 0.04
TRADING_DAYS = 252

# Etiquetas genericas — el usuario las puede renombrar en B2/C2 de la hoja Datos
LABEL_A = "Activo 1"
LABEL_B = "Activo 2"

ARIAL = "Arial"
COLOR_HDR = "1F4E78"
COLOR_HDR2 = "D9E1F2"
COLOR_INPUT = "FFF2CC"
COLOR_KPI = "E2EFDA"
GRAY = "BFBFBF"

thin = Side(border_style="thin", color=GRAY)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def style_title(cell, text, fill=COLOR_HDR, color="FFFFFF", size=14):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, size=size, color=color)
    cell.fill = PatternFill("solid", start_color=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_hdr(cell, text):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, size=10)
    cell.fill = PatternFill("solid", start_color=COLOR_HDR2)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = box


def style_input_hdr(cell, text):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, size=10, color="9C5700")
    cell.fill = PatternFill("solid", start_color=COLOR_INPUT)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = box


# =========================== Instrucciones ===========================
ws = wb.active
ws.title = "Instrucciones"

style_title(ws["A1"], "PLANTILLA DE ANALISIS COMPARATIVO", size=16)
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 28

instr = [
    ("", ""),
    ("Que hace esta plantilla", "header"),
    ("", ""),
    ("Compara DOS series de precios diarias entre si:", ""),
    ("  - Activo 1  = la serie que quieres analizar", ""),
    ("  - Activo 2  = la serie que usas como benchmark", ""),
    ("Puedes usar lo que quieras: fondos, ETFs, acciones, criptos, indices.", ""),
    ("", ""),
    ("Como usarla", "header"),
    ("", ""),
    ("1. Ve a la hoja 'Datos'.", ""),
    ("2. En B2 y C2 escribe los NOMBRES de los dos activos (opcional).", ""),
    ("3. Columna A: pega las FECHAS (dd/mm/aaaa o aaaa-mm-dd).", ""),
    ("4. Columna B: pega los PRECIOS del Activo 1.", ""),
    ("5. Columna C: pega los PRECIOS del Activo 2 (benchmark).", ""),
    ("6. Ve a 'Dashboard' y veras todas las metricas calculadas.", ""),
    ("7. Ve a 'Grafico' para la evolucion comparada (base 100) y drawdown.", ""),
    ("", ""),
    ("Parametros (modificables en Dashboard)", "header"),
    ("", ""),
    ("Tasa libre de riesgo anual: Dashboard!B4  (default 4.0%)", ""),
    ("Dias bursatiles al ano: 252", ""),
    ("", ""),
    ("Metricas que calcula", "header"),
    ("", ""),
    ("Beta             -- sensibilidad de Activo 1 respecto a Activo 2", ""),
    ("Alpha (Jensen)   -- retorno anormal sobre lo predicho por CAPM, anualizado", ""),
    ("Correlacion / R2 -- co-movimiento entre los dos activos", ""),
    ("Volatilidad      -- desviacion estandar anualizada", ""),
    ("Sharpe Ratio     -- retorno excedente / volatilidad, anualizado", ""),
    ("Tracking Error   -- desviacion estandar del diferencial Activo1 - Activo2", ""),
    ("Info Ratio       -- retorno activo / tracking error", ""),
    ("Max Drawdown     -- mayor caida pico-valle", ""),
    ("CAGR             -- retorno anualizado compuesto", ""),
    ("", ""),
    ("Notas tecnicas", "header"),
    ("", ""),
    ("- Las celdas amarillas son inputs (puedes editarlas).", ""),
    ("- Las celdas verdes son KPIs (no editar, son formulas).", ""),
    ("- La plantilla admite hasta 5,000 filas diarias (~20 anios).", ""),
    ("- Si quieres comparar contra SP500, descargalo de Yahoo (^GSPC) y pegalo en C.", ""),
    ("- Si tienes Python, el script analizador.py descarga ambos activos solo.", ""),
]

r = 2
for txt, kind in instr:
    c = ws.cell(row=r, column=1, value=txt)
    if kind == "header":
        c.font = Font(name=ARIAL, bold=True, size=12, color="1F4E78")
        c.fill = PatternFill("solid", start_color=COLOR_HDR2)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c.alignment = Alignment(vertical="center", indent=1)
    else:
        c.font = Font(name=ARIAL, size=11)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

for col, w in zip("ABCDEF", [22, 18, 18, 18, 18, 18]):
    ws.column_dimensions[col].width = w


# =========================== Datos ===========================
ws = wb.create_sheet("Datos")
style_title(ws["A1"], "DATOS (pega aqui tus precios)", size=13)
ws.merge_cells("A1:C1")
ws.row_dimensions[1].height = 22

style_input_hdr(ws["A2"], "Fecha")
style_input_hdr(ws["B2"], LABEL_A)
style_input_hdr(ws["C2"], LABEL_B)

for r in range(3, 3 + N_MAX):
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=2).number_format = "#,##0.0000"
    ws.cell(row=r, column=3).number_format = "#,##0.0000"

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16

ws["E2"] = "<- Pega tus datos. Tip: cambia los nombres en B2/C2."
ws["E2"].font = Font(name=ARIAL, italic=True, color="9C5700", size=10)
ws.column_dimensions["E"].width = 60

# Filas de ejemplo (necesarias para que charts no se rompan al abrir)
ejemplo = [
    ("2025-05-12", 100.00, 100.00),
    ("2025-05-13", 101.20, 100.40),
    ("2025-05-14", 100.80, 100.10),
]
for i, (f, a, b) in enumerate(ejemplo):
    ws.cell(row=3 + i, column=1, value=f)
    ws.cell(row=3 + i, column=2, value=a)
    ws.cell(row=3 + i, column=3, value=b)


def add_name(name, ref):
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn

add_name("FECHAS",   f"Datos!$A$3:$A${2 + N_MAX}")
add_name("PRECIO_A", f"Datos!$B$3:$B${2 + N_MAX}")
add_name("PRECIO_B", f"Datos!$C$3:$C${2 + N_MAX}")
add_name("NOMBRE_A", f"Datos!$B$2")
add_name("NOMBRE_B", f"Datos!$C$2")


# =========================== Calculos ===========================
ws = wb.create_sheet("Calculos")
style_title(ws["A1"], "CALCULOS AUXILIARES (no editar)")
ws.merge_cells("A1:L1")

headers = ["Fecha", "Px A", "Px B", "Ret A", "Ret B",
           "Exc A", "Exc B", "MaxSoFar A", "DD A", "MaxSoFar B", "DD B", "Diff (A-B)"]
for j, h in enumerate(headers, start=1):
    style_hdr(ws.cell(row=2, column=j), h)

for i in range(3, 3 + N_MAX):
    ws.cell(row=i, column=1, value=f'=IFERROR(Datos!A{i},"")')
    ws.cell(row=i, column=2, value=f'=IFERROR(Datos!B{i},"")')
    ws.cell(row=i, column=3, value=f'=IFERROR(Datos!C{i},"")')
    if i == 3:
        for col in (4, 5, 6, 7):
            ws.cell(row=i, column=col, value="")
        ws.cell(row=i, column=8, value=f'=IFERROR(IF(ISNUMBER(B{i}),B{i},""),"")')
        ws.cell(row=i, column=9, value=f'=IFERROR(IF(ISNUMBER(B{i}),0,""),"")')
        ws.cell(row=i, column=10, value=f'=IFERROR(IF(ISNUMBER(C{i}),C{i},""),"")')
        ws.cell(row=i, column=11, value=f'=IFERROR(IF(ISNUMBER(C{i}),0,""),"")')
        ws.cell(row=i, column=12, value="")
    else:
        ws.cell(row=i, column=4,
                value=f'=IFERROR(IF(AND(ISNUMBER(B{i}),ISNUMBER(B{i-1}),B{i-1}<>0),B{i}/B{i-1}-1,""),"")')
        ws.cell(row=i, column=5,
                value=f'=IFERROR(IF(AND(ISNUMBER(C{i}),ISNUMBER(C{i-1}),C{i-1}<>0),C{i}/C{i-1}-1,""),"")')
        ws.cell(row=i, column=6, value=f'=IFERROR(D{i}-Dashboard!$B$4/252,"")')
        ws.cell(row=i, column=7, value=f'=IFERROR(E{i}-Dashboard!$B$4/252,"")')
        ws.cell(row=i, column=8,
                value=f'=IFERROR(IF(ISNUMBER(B{i}),MAX(B{i},IFERROR(H{i-1},B{i})),""),"")')
        ws.cell(row=i, column=9,
                value=f'=IFERROR(IF(AND(ISNUMBER(B{i}),ISNUMBER(H{i})),B{i}/H{i}-1,""),"")')
        ws.cell(row=i, column=10,
                value=f'=IFERROR(IF(ISNUMBER(C{i}),MAX(C{i},IFERROR(J{i-1},C{i})),""),"")')
        ws.cell(row=i, column=11,
                value=f'=IFERROR(IF(AND(ISNUMBER(C{i}),ISNUMBER(J{i})),C{i}/J{i}-1,""),"")')
        ws.cell(row=i, column=12,
                value=f'=IFERROR(IF(AND(ISNUMBER(D{i}),ISNUMBER(E{i})),D{i}-E{i},""),"")')

# Formatos
for i in range(3, 3 + N_MAX):
    ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=i, column=2).number_format = "#,##0.0000"
    ws.cell(row=i, column=3).number_format = "#,##0.0000"
    for col in ("D", "E", "F", "G", "L"):
        ws[f"{col}{i}"].number_format = "0.0000%"
    ws.cell(row=i, column=8).number_format = "#,##0.0000"
    ws.cell(row=i, column=9).number_format = "0.00%"
    ws.cell(row=i, column=10).number_format = "#,##0.0000"
    ws.cell(row=i, column=11).number_format = "0.00%"

for col, w in zip("ABCDEFGHIJKL", [12, 12, 12, 13, 13, 13, 13, 13, 12, 13, 12, 13]):
    ws.column_dimensions[col].width = w

add_name("RET_A",  f"Calculos!$D$3:$D${2 + N_MAX}")
add_name("RET_B",  f"Calculos!$E$3:$E${2 + N_MAX}")
add_name("EXC_A",  f"Calculos!$F$3:$F${2 + N_MAX}")
add_name("EXC_B",  f"Calculos!$G$3:$G${2 + N_MAX}")
add_name("DD_A",   f"Calculos!$I$3:$I${2 + N_MAX}")
add_name("DD_B",   f"Calculos!$K$3:$K${2 + N_MAX}")
add_name("DIFF_AB",f"Calculos!$L$3:$L${2 + N_MAX}")


# =========================== Dashboard ===========================
ws = wb.create_sheet("Dashboard")
style_title(ws["A1"], "DASHBOARD COMPARATIVO  -  ACTIVO 1  vs  ACTIVO 2", size=14)
ws.merge_cells("A1:D1")
ws.row_dimensions[1].height = 26

ws["A3"] = "Parametros"
ws["A3"].font = Font(name=ARIAL, bold=True, size=11, color="1F4E78")

ws["A4"] = "Tasa libre de riesgo anual"
ws["A4"].font = Font(name=ARIAL, size=10)
ws["B4"] = RF_DEFAULT
ws["B4"].font = Font(name=ARIAL, bold=True, color="0000FF", size=11)
ws["B4"].fill = PatternFill("solid", start_color=COLOR_INPUT)
ws["B4"].number_format = "0.00%"
ws["B4"].alignment = Alignment(horizontal="center")
ws["B4"].border = box

ws["A5"] = "Dias bursatiles / ano"
ws["A5"].font = Font(name=ARIAL, size=10)
ws["B5"] = TRADING_DAYS
ws["B5"].font = Font(name=ARIAL, bold=True, color="0000FF", size=11)
ws["B5"].fill = PatternFill("solid", start_color=COLOR_INPUT)
ws["B5"].alignment = Alignment(horizontal="center")
ws["B5"].border = box

ws["A6"] = "Observaciones (n)"
ws["A6"].font = Font(name=ARIAL, size=10)
ws["B6"] = "=COUNT(RET_A)"
ws["B6"].font = Font(name=ARIAL, bold=True)
ws["B6"].fill = PatternFill("solid", start_color=COLOR_KPI)
ws["B6"].number_format = "#,##0"
ws["B6"].alignment = Alignment(horizontal="center")
ws["B6"].border = box

ws["A8"] = "METRICA"
ws["B8"] = "=NOMBRE_A"
ws["C8"] = "=NOMBRE_B"
ws["D8"] = "INTERPRETACION"
for col in "ABCD":
    cell = ws[f"{col}8"]
    cell.font = Font(name=ARIAL, bold=True, size=10)
    cell.fill = PatternFill("solid", start_color=COLOR_HDR2)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = box

metricas = [
    ("Retorno medio diario",
     "=IFERROR(AVERAGE(RET_A),NA())",
     "=IFERROR(AVERAGE(RET_B),NA())",
     "0.0000%", "Promedio simple de retornos diarios"),
    ("CAGR (anualizado compuesto)",
     '=IFERROR((LOOKUP(2,1/ISNUMBER(PRECIO_A),PRECIO_A)/INDEX(PRECIO_A,MATCH(TRUE,ISNUMBER(PRECIO_A),0)))^($B$5/COUNT(RET_A))-1,NA())',
     '=IFERROR((LOOKUP(2,1/ISNUMBER(PRECIO_B),PRECIO_B)/INDEX(PRECIO_B,MATCH(TRUE,ISNUMBER(PRECIO_B),0)))^($B$5/COUNT(RET_B))-1,NA())',
     "0.00%", "Tasa anual compuesta equivalente"),
    ("Volatilidad anualizada",
     "=IFERROR(STDEV(RET_A)*SQRT($B$5),NA())",
     "=IFERROR(STDEV(RET_B)*SQRT($B$5),NA())",
     "0.00%", "Desviacion estandar anualizada"),
    ("Sharpe Ratio (anualizado)",
     "=IFERROR((AVERAGE(EXC_A)*$B$5)/(STDEV(RET_A)*SQRT($B$5)),NA())",
     "=IFERROR((AVERAGE(EXC_B)*$B$5)/(STDEV(RET_B)*SQRT($B$5)),NA())",
     "0.000", "Retorno excedente / volatilidad"),
    ("Max Drawdown",
     "=IFERROR(MIN(DD_A),NA())",
     "=IFERROR(MIN(DD_B),NA())",
     "0.00%", "Mayor caida pico-valle"),
    ("Beta (Activo 1 vs Activo 2)",
     "=IFERROR(SLOPE(EXC_A,EXC_B),NA())", "=1",
     "0.000", "<1 menos volatil ; >1 mas volatil ; =1 igual al benchmark"),
    ("Alpha Jensen (anualizado)",
     "=IFERROR((AVERAGE(EXC_A)-SLOPE(EXC_A,EXC_B)*AVERAGE(EXC_B))*$B$5,NA())", "=0",
     "0.00%", "Retorno excedente sobre lo predicho por CAPM"),
    ("Correlacion (Pearson)",
     "=IFERROR(CORREL(RET_A,RET_B),NA())", "=1",
     "0.000", "Co-movimiento (-1 a +1)"),
    ("R-cuadrado",
     "=IFERROR(CORREL(RET_A,RET_B)^2,NA())", "=1",
     "0.00%", "Varianza explicada por Activo 2"),
    ("Tracking Error (anualizado)",
     "=IFERROR(STDEV(DIFF_AB)*SQRT($B$5),NA())", "=0",
     "0.00%", "Volatilidad del diferencial"),
    ("Information Ratio",
     "=IFERROR((AVERAGE(DIFF_AB)*$B$5)/(STDEV(DIFF_AB)*SQRT($B$5)),NA())", '=NA()',
     "0.000", "Retorno activo / Tracking Error"),
]

start_row = 9
for i, (lbl, f_a, f_b, fmt, interp) in enumerate(metricas):
    r = start_row + i
    ws.cell(row=r, column=1, value=lbl).font = Font(name=ARIAL, bold=True, size=10)
    ws.cell(row=r, column=1).border = box

    cs = ws.cell(row=r, column=2, value=f_a)
    cs.number_format = fmt
    cs.font = Font(name=ARIAL, bold=True, size=10)
    cs.alignment = Alignment(horizontal="center")
    cs.fill = PatternFill("solid", start_color=COLOR_KPI)
    cs.border = box

    cb = ws.cell(row=r, column=3, value=f_b)
    cb.number_format = fmt
    cb.font = Font(name=ARIAL, size=10)
    cb.alignment = Alignment(horizontal="center")
    cb.fill = PatternFill("solid", start_color="F2F2F2")
    cb.border = box

    ci = ws.cell(row=r, column=4, value=interp)
    ci.font = Font(name=ARIAL, italic=True, size=10, color="595959")
    ci.border = box
    ci.alignment = Alignment(horizontal="left", indent=1, vertical="center", wrap_text=True)

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 56

beta_row = start_row + 5
ws.conditional_formatting.add(
    f"B{beta_row}",
    CellIsRule(operator="greaterThan", formula=["1.1"],
               fill=PatternFill("solid", start_color="F8CBAD")))
ws.conditional_formatting.add(
    f"B{beta_row}",
    CellIsRule(operator="lessThan", formula=["0.9"],
               fill=PatternFill("solid", start_color="C6EFCE")))


# =========================== Grafico ===========================
ws_g = wb.create_sheet("Grafico")
style_title(ws_g["A1"], "EVOLUCION COMPARADA  (Base 100)")
ws_g.merge_cells("A1:E1")

# Headers dinamicos: usan los nombres de B2/C2 de Datos
style_hdr(ws_g["A2"], "Fecha")
ws_g["B2"] = '=NOMBRE_A&" (base 100)"'
ws_g["C2"] = '=NOMBRE_B&" (base 100)"'
for c in ["B2", "C2"]:
    ws_g[c].font = Font(name=ARIAL, bold=True, size=10)
    ws_g[c].fill = PatternFill("solid", start_color=COLOR_HDR2)
    ws_g[c].alignment = Alignment(horizontal="center")
    ws_g[c].border = box

ws_g["E2"] = "Primer Px Activo 1"
ws_g["F2"] = '=INDEX(PRECIO_A,MATCH(TRUE,ISNUMBER(PRECIO_A),0))'
ws_g["E3"] = "Primer Px Activo 2"
ws_g["F3"] = '=INDEX(PRECIO_B,MATCH(TRUE,ISNUMBER(PRECIO_B),0))'
for c in ["F2", "F3"]:
    ws_g[c].number_format = "#,##0.0000"
    ws_g[c].font = Font(name=ARIAL, bold=True, size=10)
    ws_g[c].fill = PatternFill("solid", start_color=COLOR_KPI)
for c in ["E2", "E3"]:
    ws_g[c].font = Font(name=ARIAL, italic=True, size=10, color="595959")

for i in range(3, 3 + N_MAX):
    ws_g.cell(row=i, column=1, value=f'=IFERROR(Datos!A{i},"")')
    ws_g.cell(row=i, column=1).number_format = "yyyy-mm-dd"
    ws_g.cell(row=i, column=2,
              value=f'=IFERROR(IF(ISNUMBER(Datos!B{i}),Datos!B{i}/$F$2*100,NA()),NA())')
    ws_g.cell(row=i, column=2).number_format = "#,##0.00"
    ws_g.cell(row=i, column=3,
              value=f'=IFERROR(IF(ISNUMBER(Datos!C{i}),Datos!C{i}/$F$3*100,NA()),NA())')
    ws_g.cell(row=i, column=3).number_format = "#,##0.00"

for col, w in zip("ABCDEF", [13, 22, 22, 4, 22, 14]):
    ws_g.column_dimensions[col].width = w

chart = LineChart()
chart.title = "Activo 1 vs Activo 2 (Base 100)"
chart.y_axis.title = "Nivel (Base 100)"
chart.x_axis.title = "Fecha"
chart.height = 14
chart.width = 26
chart.style = 12

data = Reference(ws_g, min_col=2, max_col=3, min_row=2, max_row=2 + N_MAX)
cats = Reference(ws_g, min_col=1, min_row=3, max_row=2 + N_MAX)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_g.add_chart(chart, "H2")

chart2 = LineChart()
chart2.title = "Drawdown (%) - Activo 1 vs Activo 2"
chart2.y_axis.title = "Drawdown"
chart2.x_axis.title = "Fecha"
chart2.height = 12
chart2.width = 26
data2 = Reference(wb["Calculos"], min_col=9, max_col=9, min_row=2, max_row=2 + N_MAX)
data2b = Reference(wb["Calculos"], min_col=11, max_col=11, min_row=2, max_row=2 + N_MAX)
cats2 = Reference(wb["Calculos"], min_col=1, min_row=3, max_row=2 + N_MAX)
chart2.add_data(data2, titles_from_data=True)
chart2.add_data(data2b, titles_from_data=True)
chart2.set_categories(cats2)
ws_g.add_chart(chart2, "H32")


wb["Instrucciones"].sheet_view.showGridLines = False
wb["Dashboard"].sheet_view.showGridLines = False
wb["Grafico"].sheet_view.showGridLines = False
wb["Instrucciones"].sheet_properties.tabColor = "1F4E78"
wb["Datos"].sheet_properties.tabColor = "FFC000"
wb["Calculos"].sheet_properties.tabColor = "808080"
wb["Dashboard"].sheet_properties.tabColor = "00B050"
wb["Grafico"].sheet_properties.tabColor = "4472C4"

order = ["Instrucciones", "Datos", "Dashboard", "Grafico", "Calculos"]
wb._sheets = [wb[s] for s in order]

wb.save(OUT)
print(f"Plantilla guardada: {OUT}")
      