"""
Analizador Comparativo de Series de Tiempo
==========================================
Herramienta de análisis técnico que compara una o varias series de precios
contra un BENCHMARK configurable (default S&P 500 ^GSPC, pero puede ser
cualquier ticker Yahoo: GLD, QQQ, ^MXX, AAPL, BTC-USD, etc.).

Calcula métricas estadísticas clave:

    - Beta (sensibilidad al mercado)
    - Alpha (Jensen, anualizado)
    - Correlación de Pearson y R²
    - Volatilidad anualizada
    - Sharpe Ratio
    - Tracking Error
    - Maximum Drawdown
    - Retornos acumulados

Genera:
    1) Reporte Excel con todas las métricas y series alineadas
    2) Gráfico HTML interactivo (Plotly) con dashboard completo

Uso:
    python analizador.py <archivo_de_datos> [--rf 0.045] [--start 2020-01-01] [--benchmark ^GSPC]

Formatos aceptados: .xlsx, .csv, .txt (delimitado por tabs/comas/punto y coma)
Estructura esperada del archivo:
    - Columna 1: Fecha (cualquier formato común)
    - Columnas 2..N: Precios o niveles de la(s) serie(s) a analizar
"""

from __future__ import annotations
import argparse
import os
import sys
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# 1. CARGA Y LIMPIEZA DE DATOS
# ---------------------------------------------------------------------------
def cargar_serie(ruta: str) -> pd.DataFrame:
    """Carga la serie del usuario desde xlsx, csv o txt.

    Detecta separador automáticamente y normaliza la columna de fecha.
    Devuelve un DataFrame con índice DatetimeIndex y columnas numéricas.
    """
    ruta_p = Path(ruta)
    if not ruta_p.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")

    ext = ruta_p.suffix.lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        df = pd.read_excel(ruta_p)
    elif ext in (".csv", ".txt", ".tsv"):
        # Intento detectar separador
        df = pd.read_csv(ruta_p, sep=None, engine="python")
    else:
        raise ValueError(f"Extensión no soportada: {ext}")

    if df.shape[1] < 2:
        raise ValueError("El archivo debe tener al menos 2 columnas: Fecha y Valor")

    # Primera columna como fecha
    df.columns = [str(c).strip() for c in df.columns]
    fecha_col = df.columns[0]
    df[fecha_col] = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=False)
    df = df.dropna(subset=[fecha_col]).set_index(fecha_col)
    df.index.name = "Fecha"

    # Convertir las demás columnas a numérico
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df.sort_index().dropna(how="all")
    # Quita filas con todos NaN, conserva las parciales
    return df


def cargar_benchmark(ticker: str, start: str, end: str | None) -> pd.Series:
    """Descarga el benchmark (default S&P 500) desde Yahoo Finance."""
    try:
        import yfinance as yf
    except ImportError as e:
        raise ImportError(
            "yfinance no está instalado. Ejecuta: pip install yfinance"
        ) from e

    print(f"  -> Descargando {ticker} desde {start} hasta {end or 'hoy'}...")
    data = yf.download(ticker, start=start, end=end, progress=False, auto_adjust=True)
    if data.empty:
        raise RuntimeError(f"No se pudo descargar datos para {ticker}")
    # Yahoo a veces regresa MultiIndex
    if isinstance(data.columns, pd.MultiIndex):
        data.columns = [c[0] for c in data.columns]
    serie = data["Close"].rename(ticker)
    serie.index = pd.to_datetime(serie.index)
    return serie


# ---------------------------------------------------------------------------
# 2. CÁLCULO DE MÉTRICAS
# ---------------------------------------------------------------------------
PERIODOS_ANIO = 252  # días hábiles bursátiles


def retornos_log(precios: pd.Series) -> pd.Series:
    return np.log(precios / precios.shift(1)).dropna()


def retornos_simples(precios: pd.Series) -> pd.Series:
    return precios.pct_change().dropna()


def calcular_metricas(
    serie: pd.Series,
    benchmark: pd.Series,
    rf_anual: float = 0.0,
) -> dict:
    """Calcula todas las métricas para una serie contra el benchmark.

    Asume que ambas series ya están alineadas (mismas fechas).
    """
    r_a = retornos_simples(serie)
    r_b = retornos_simples(benchmark)
    # Alinea por intersección
    idx = r_a.index.intersection(r_b.index)
    r_a, r_b = r_a.loc[idx], r_b.loc[idx]

    rf_diario = (1 + rf_anual) ** (1 / PERIODOS_ANIO) - 1

    # --- Beta y Alpha (regresión OLS de r_a sobre r_b) ---
    cov = np.cov(r_a, r_b, ddof=1)
    var_b = cov[1, 1]
    beta = cov[0, 1] / var_b if var_b > 0 else np.nan

    # Alpha de Jensen anualizado: alpha = (mean(r_a) - rf) - beta * (mean(r_b) - rf)
    mean_a, mean_b = r_a.mean(), r_b.mean()
    alpha_diario = (mean_a - rf_diario) - beta * (mean_b - rf_diario)
    alpha_anual = alpha_diario * PERIODOS_ANIO

    # --- Correlación y R² ---
    corr = r_a.corr(r_b)
    r2 = corr ** 2

    # --- Volatilidad anualizada ---
    vol_a = r_a.std(ddof=1) * np.sqrt(PERIODOS_ANIO)
    vol_b = r_b.std(ddof=1) * np.sqrt(PERIODOS_ANIO)

    # --- Sharpe ---
    sharpe_a = (mean_a - rf_diario) / r_a.std(ddof=1) * np.sqrt(PERIODOS_ANIO) if r_a.std(ddof=1) > 0 else np.nan
    sharpe_b = (mean_b - rf_diario) / r_b.std(ddof=1) * np.sqrt(PERIODOS_ANIO) if r_b.std(ddof=1) > 0 else np.nan

    # --- Tracking Error (vol del diferencial de retornos, anualizada) ---
    diff = r_a - r_b
    tracking_error = diff.std(ddof=1) * np.sqrt(PERIODOS_ANIO)

    # --- Information Ratio ---
    info_ratio = (diff.mean() * PERIODOS_ANIO) / tracking_error if tracking_error > 0 else np.nan

    # --- Retornos acumulados ---
    nav_a = (1 + r_a).prod() - 1
    nav_b = (1 + r_b).prod() - 1
    n_anios = len(r_a) / PERIODOS_ANIO
    cagr_a = (1 + nav_a) ** (1 / n_anios) - 1 if n_anios > 0 else np.nan
    cagr_b = (1 + nav_b) ** (1 / n_anios) - 1 if n_anios > 0 else np.nan

    # --- Max Drawdown ---
    def max_dd(precios):
        cum = precios / precios.iloc[0]
        peak = cum.cummax()
        dd = (cum / peak) - 1
        return dd.min()

    mdd_a = max_dd(serie.loc[idx[0]:idx[-1]])
    mdd_b = max_dd(benchmark.loc[idx[0]:idx[-1]])

    return {
        "Observaciones": int(len(r_a)),
        "Periodo_inicio": str(idx[0].date()),
        "Periodo_fin": str(idx[-1].date()),
        "Beta": round(beta, 4),
        "Alpha_anual (Jensen)": round(alpha_anual, 4),
        "Correlacion": round(corr, 4),
        "R_cuadrado": round(r2, 4),
        "Volatilidad_anual_serie": round(vol_a, 4),
        "Volatilidad_anual_benchmark": round(vol_b, 4),
        "Sharpe_serie": round(sharpe_a, 4),
        "Sharpe_benchmark": round(sharpe_b, 4),
        "Tracking_Error_anual": round(tracking_error, 4),
        "Information_Ratio": round(info_ratio, 4),
        "Retorno_acumulado_serie": round(nav_a, 4),
        "Retorno_acumulado_benchmark": round(nav_b, 4),
        "CAGR_serie": round(cagr_a, 4),
        "CAGR_benchmark": round(cagr_b, 4),
        "Max_Drawdown_serie": round(mdd_a, 4),
        "Max_Drawdown_benchmark": round(mdd_b, 4),
        "Tasa_libre_riesgo_usada": rf_anual,
    }


# ---------------------------------------------------------------------------
# 3. ALINEACIÓN
# ---------------------------------------------------------------------------
def alinear(df_user: pd.DataFrame, bench: pd.Series) -> pd.DataFrame:
    """Une la(s) serie(s) del usuario con el benchmark por fecha (intersección)."""
    combinado = df_user.join(bench, how="inner")
    combinado = combinado.ffill().dropna(how="any")
    return combinado


# ---------------------------------------------------------------------------
# 4. EXPORTACIÓN
# ---------------------------------------------------------------------------
def exportar_excel(
    combinado: pd.DataFrame,
    resultados: dict,
    salida: str,
) -> None:
    """Genera reporte Excel con: Resumen, Datos alineados, Retornos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference

    wb = Workbook()

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "ANÁLISIS DE SERIES vs S&P 500"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color="1F4E78")
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    fila = 3
    encabezado_fill = PatternFill("solid", start_color="D9E1F2")
    border_thin = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for nombre_serie, metr in resultados.items():
        ws.cell(row=fila, column=1, value=f"Serie: {nombre_serie}").font = Font(bold=True, size=12)
        ws.cell(row=fila, column=1).fill = PatternFill("solid", start_color="FCE4D6")
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        fila += 1
        ws.cell(row=fila, column=1, value="Métrica").font = Font(bold=True)
        ws.cell(row=fila, column=2, value="Valor").font = Font(bold=True)
        ws.cell(row=fila, column=3, value="Interpretación").font = Font(bold=True)
        for c in range(1, 4):
            ws.cell(row=fila, column=c).fill = encabezado_fill
            ws.cell(row=fila, column=c).border = border_thin
        fila += 1
        interpretaciones = _interpretar(metr)
        for k, v in metr.items():
            ws.cell(row=fila, column=1, value=k)
            ws.cell(row=fila, column=2, value=v)
            ws.cell(row=fila, column=3, value=interpretaciones.get(k, ""))
            for c in range(1, 4):
                ws.cell(row=fila, column=c).border = border_thin
            # Formato numérico
            if isinstance(v, float):
                if any(x in k for x in ["Volatilidad", "Retorno", "Alpha", "CAGR", "Drawdown", "Tracking", "Tasa"]):
                    ws.cell(row=fila, column=2).number_format = "0.00%"
                else:
                    ws.cell(row=fila, column=2).number_format = "0.0000"
            fila += 1
        fila += 2

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 60

    # --- Hoja 2: Datos alineados (precios) ---
    ws2 = wb.create_sheet("Precios_alineados")
    ws2.append(["Fecha"] + list(combinado.columns))
    for r in dataframe_to_rows(combinado, index=True, header=False):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = encabezado_fill
    ws2.column_dimensions["A"].width = 12
    for col_idx in range(2, combinado.shape[1] + 2):
        col_letter = ws2.cell(row=1, column=col_idx).column_letter
        ws2.column_dimensions[col_letter].width = 14

    # --- Hoja 3: Retornos diarios ---
    ws3 = wb.create_sheet("Retornos_diarios")
    rets = combinado.pct_change().dropna(how="all")
    ws3.append(["Fecha"] + list(rets.columns))
    for r in dataframe_to_rows(rets, index=True, header=False):
        ws3.append(r)
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = encabezado_fill
    # formato porcentual
    for col_idx in range(2, rets.shape[1] + 2):
        col_letter = ws3.cell(row=1, column=col_idx).column_letter
        for r in range(2, ws3.max_row + 1):
            ws3.cell(row=r, column=col_idx).number_format = "0.00%"
        ws3.column_dimensions[col_letter].width = 12

    # --- Hoja 4: NAV base 100 con gráfico ---
    ws4 = wb.create_sheet("NAV_base100")
    nav = combinado.divide(combinado.iloc[0]).multiply(100)
    ws4.append(["Fecha"] + list(nav.columns))
    for r in dataframe_to_rows(nav, index=True, header=False):
        ws4.append(r)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.fill = encabezado_fill
    ws4.column_dimensions["A"].width = 12

    # Gráfico
    chart = LineChart()
    chart.title = "Evolución comparada (Base 100)"
    chart.y_axis.title = "Nivel (Base 100)"
    chart.x_axis.title = "Fecha"
    chart.height = 12
    chart.width = 22
    data_ref = Reference(ws4, min_col=2, max_col=nav.shape[1] + 1, min_row=1, max_row=ws4.max_row)
    cats_ref = Reference(ws4, min_col=1, max_col=1, min_row=2, max_row=ws4.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws4.add_chart(chart, "G2")

    wb.save(salida)
    print(f"  -> Reporte Excel: {salida}")


def _interpretar(m: dict) -> dict:
    """Genera frases cortas de interpretación para cada métrica."""
    beta = m.get("Beta", np.nan)
    alpha = m.get("Alpha_anual (Jensen)", np.nan)
    corr = m.get("Correlacion", np.nan)
    sharpe = m.get("Sharpe_serie", np.nan)
    return {
        "Beta": (
            f"Sensibilidad al mercado. {'Más volátil que el mercado.' if beta > 1.1 else 'Menos volátil que el mercado.' if beta < 0.9 else 'Comportamiento similar al mercado.'}"
        ),
        "Alpha_anual (Jensen)": (
            f"{'Sobre-rendimiento ajustado por riesgo.' if alpha > 0 else 'Sub-rendimiento ajustado por riesgo.'}"
        ),
        "Correlacion": (
            f"{'Alta correlación' if abs(corr) > 0.7 else 'Correlación moderada' if abs(corr) > 0.4 else 'Baja correlación'} con el benchmark."
        ),
        "R_cuadrado": "% del movimiento explicado por el benchmark.",
        "Sharpe_serie": (
            f"{'Excelente' if sharpe > 1.5 else 'Bueno' if sharpe > 1 else 'Aceptable' if sharpe > 0.5 else 'Pobre'} retorno por unidad de riesgo."
        ),
        "Tracking_Error_anual": "Desviación anual respecto al benchmark.",
        "Information_Ratio": "Retorno activo por unidad de tracking error.",
        "Max_Drawdown_serie": "Mayor caída pico-valle observada.",
    }


# ---------------------------------------------------------------------------
# 5. GRÁFICO HTML INTERACTIVO
# ---------------------------------------------------------------------------
def exportar_html(combinado: pd.DataFrame, resultados: dict, benchmark_col: str, salida: str) -> None:
    """Genera dashboard HTML interactivo con Plotly."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print("  ! plotly no instalado; omito gráfico HTML. Instala con: pip install plotly")
        return

    nav = combinado.divide(combinado.iloc[0]).multiply(100)
    rets = combinado.pct_change().dropna()

    series_user = [c for c in combinado.columns if c != benchmark_col]
    n_series = len(series_user)

    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=(
            "Evolución Base 100",
            "Drawdown (%)",
            "Retornos diarios — dispersión vs benchmark",
            "Volatilidad rolling 60d (anualizada)",
            "Correlación rolling 60d",
            "Beta rolling 60d",
        ),
        specs=[[{}, {}], [{}, {}], [{}, {}]],
        vertical_spacing=0.10,
        horizontal_spacing=0.10,
    )

    # 1. NAV base 100
    for c in combinado.columns:
        fig.add_trace(go.Scatter(x=nav.index, y=nav[c], name=c, mode="lines"), row=1, col=1)

    # 2. Drawdown
    for c in combinado.columns:
        cum = combinado[c] / combinado[c].iloc[0]
        dd = (cum / cum.cummax() - 1) * 100
        fig.add_trace(go.Scatter(x=dd.index, y=dd, name=f"DD {c}", mode="lines",
                                  fill="tozeroy", opacity=0.5, showlegend=False), row=1, col=2)

    # 3. Scatter rets serie vs benchmark
    for c in series_user:
        fig.add_trace(
            go.Scatter(x=rets[benchmark_col] * 100, y=rets[c] * 100,
                       mode="markers", name=f"Rets {c}",
                       marker=dict(size=4, opacity=0.5), showlegend=False),
            row=2, col=1,
        )

    # 4. Volatilidad rolling
    vol = rets.rolling(60).std() * np.sqrt(PERIODOS_ANIO) * 100
    for c in combinado.columns:
        fig.add_trace(go.Scatter(x=vol.index, y=vol[c], name=f"Vol {c}", mode="lines",
                                  showlegend=False), row=2, col=2)

    # 5. Correlación rolling
    for c in series_user:
        cr = rets[c].rolling(60).corr(rets[benchmark_col])
        fig.add_trace(go.Scatter(x=cr.index, y=cr, name=f"Corr {c}", mode="lines",
                                  showlegend=False), row=3, col=1)

    # 6. Beta rolling = Cov/Var en ventana 60d
    for c in series_user:
        cov_r = rets[c].rolling(60).cov(rets[benchmark_col])
        var_r = rets[benchmark_col].rolling(60).var()
        beta_r = cov_r / var_r
        fig.add_trace(go.Scatter(x=beta_r.index, y=beta_r, name=f"Beta {c}", mode="lines",
                                  showlegend=False), row=3, col=2)

    fig.update_layout(
        title=dict(text="Dashboard de análisis técnico vs benchmark", font=dict(size=20)),
        height=1200,
        template="plotly_white",
        hovermode="x unified",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
    )
    fig.update_xaxes(title_text="Retorno benchmark (%)", row=2, col=1)
    fig.update_yaxes(title_text="Retorno serie (%)", row=2, col=1)

    # Tabla de métricas anexada al HTML
    tabla_html = "<h2 style='font-family:Arial'>Métricas clave</h2><table style='border-collapse:collapse;font-family:Arial'>"
    headers = list(next(iter(resultados.values())).keys())
    tabla_html += "<tr style='background:#1F4E78;color:white'><th style='padding:6px;border:1px solid #ccc'>Métrica</th>"
    for s in resultados:
        tabla_html += f"<th style='padding:6px;border:1px solid #ccc'>{s}</th>"
    tabla_html += "</tr>"
    for h in headers:
        tabla_html += f"<tr><td style='padding:6px;border:1px solid #ccc;background:#f2f2f2'><b>{h}</b></td>"
        for s, m in resultados.items():
            tabla_html += f"<td style='padding:6px;border:1px solid #ccc;text-align:right'>{m[h]}</td>"
        tabla_html += "</tr>"
    tabla_html += "</table>"

    html_chart = fig.to_html(include_plotlyjs="cdn", full_html=False)
    full = f"""<!doctype html><html><head><meta charset='utf-8'>
<title>Análisis vs S&P 500</title></head><body style='font-family:Arial;margin:20px'>
<h1>Análisis técnico de series vs benchmark</h1>
{html_chart}
{tabla_html}
<p style='color:#888;font-size:11px'>Generado el {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
</body></html>"""

    with open(salida, "w", encoding="utf-8") as f:
        f.write(full)
    print(f"  -> Dashboard HTML: {salida}")


# ---------------------------------------------------------------------------
# 6. MAIN
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 5b. RESOLUCIÓN DE FUENTES (archivo local o ticker Yahoo)
# ---------------------------------------------------------------------------
EXT_ARCHIVO = (".xlsx", ".xls", ".xlsm", ".csv", ".txt", ".tsv")

def es_archivo(fuente: str) -> bool:
    """True si la 'fuente' parece un archivo (tiene extensión soportada
    o contiene separador de ruta)."""
    s = fuente.lower()
    if s.endswith(EXT_ARCHIVO):
        return True
    if any(sep in fuente for sep in ("/", "\\")):
        return True
    return False


def cargar_fuente(fuente: str, start: str | None, end: str | None) -> pd.DataFrame:
    """Carga una 'fuente' que puede ser un archivo local o un ticker Yahoo.

    Devuelve siempre un DataFrame indexado por Fecha con 1+ columnas numéricas.
    """
    if es_archivo(fuente):
        print(f"  -> Cargando archivo: {fuente}")
        return cargar_serie(fuente)
    # Trata como ticker Yahoo
    print(f"  -> Descargando ticker Yahoo: {fuente}")
    serie = cargar_benchmark(fuente, start, end)
    return serie.to_frame()


# ---------------------------------------------------------------------------
# 6. MAIN
# ---------------------------------------------------------------------------
DEFAULT_LOOKBACK_ANIOS = 10  # default cuando no se especifica --start

def main():
    ap = argparse.ArgumentParser(
        description="Compara series de tiempo contra un benchmark configurable. "
                    "Acepta archivos locales (.xlsx/.csv/.txt) y/o tickers Yahoo "
                    "(GLD, AAPL, ^MXX, BTC-USD, ^GSPC, etc.). El benchmark por "
                    "defecto es el S&P 500 pero puedes cambiarlo con --benchmark."
    )
    ap.add_argument("fuentes", nargs="+",
                    help="Una o más fuentes: rutas a archivos y/o tickers Yahoo")
    ap.add_argument("--benchmark", default="^GSPC",
                    help="Ticker Yahoo del benchmark (default: ^GSPC = S&P 500)")
    ap.add_argument("--rf", type=float, default=0.045,
                    help="Tasa libre de riesgo anual (default 4.5%%)")
    ap.add_argument("--start", default=None,
                    help=f"Fecha inicio YYYY-MM-DD (default: hoy − {DEFAULT_LOOKBACK_ANIOS} años "
                         "si todas las fuentes son tickers; si hay archivos, usa su primera fecha)")
    ap.add_argument("--end", default=None,
                    help="Fecha fin YYYY-MM-DD (default: hoy)")
    ap.add_argument("--out", default=None,
                    help="Carpeta de salida (default: cwd)")
    args = ap.parse_args()

    print("\n=== ANALIZADOR COMPARATIVO DE SERIES ===")
    print(f"Fuentes: {args.fuentes}")
    print(f"Benchmark: {args.benchmark}")
    print(f"Tasa libre de riesgo: {args.rf*100:.2f}%")

    # Si todas las fuentes son tickers, usamos lookback default de 10 años
    todas_tickers = all(not es_archivo(f) for f in args.fuentes)
    if args.start is None and todas_tickers:
        start = (datetime.now() - pd.DateOffset(years=DEFAULT_LOOKBACK_ANIOS)).strftime("%Y-%m-%d")
        print(f"  -> Sin --start; usando ventana de {DEFAULT_LOOKBACK_ANIOS} años: desde {start}")
    else:
        start = args.start
    end = args.end or datetime.now().strftime("%Y-%m-%d")

    # Carga cada fuente como DataFrame
    dataframes = []
    for f in args.fuentes:
        df_f = cargar_fuente(f, start, end)
        dataframes.append(df_f)
        print(f"     {df_f.shape[0]} filas, columnas: {list(df_f.columns)}")

    # Si había archivos y no se dió --start, ajustar a la primera fecha real
    if start is None:
        start = min(df.index.min() for df in dataframes).strftime("%Y-%m-%d")

    # Combina todas las series del usuario
    df_user = pd.concat(dataframes, axis=1, join="outer")

    # Descarga benchmark
    bench = cargar_benchmark(args.benchmark, start, end)

    # Alinea por intersección
    combinado = alinear(df_user, bench)
    print(f"  -> Datos alineados: {combinado.shape[0]} filas comunes "
          f"({combinado.index.min().date()} → {combinado.index.max().date()})")

    benchmark_col = bench.name
    resultados = {}
    for c in [col for col in combinado.columns if col != benchmark_col]:
        resultados[c] = calcular_metricas(combinado[c], combinado[benchmark_col], args.rf)
        print(f"  -> {c}: Beta={resultados[c]['Beta']:.3f}  "
              f"R²={resultados[c]['R_cuadrado']:.3f}  "
              f"Sharpe={resultados[c]['Sharpe_serie']:.3f}  "
              f"Vol={resultados[c]['Volatilidad_anual_serie']*100:.1f}%")

    # Directorio de salida
    if args.out:
        out_dir = Path(args.out)
    elif es_archivo(args.fuentes[0]):
        out_dir = Path(args.fuentes[0]).parent
    else:
        out_dir = Path.cwd()
    out_dir.mkdir(parents=True, exist_ok=True)

    # Nombre base = primer ticker/archivo (sin extensión ni símbolos)
    primera = args.fuentes[0]
    base = Path(primera).stem if es_archivo(primera) else primera.replace("^", "")
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    excel_out = out_dir / f"{base}_analisis_{stamp}.xlsx"
    html_out = out_dir / f"{base}_dashboard_{stamp}.html"

    exportar_excel(combinado, resultados, str(excel_out))
    exportar_html(combinado, resultados, benchmark_col, str(html_out))

    print("\n[OK] Listo. Abre el dashboard HTML y el reporte Excel.")


if __name__ == "__main__":
    main()
